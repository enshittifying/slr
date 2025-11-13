"""
Spreadsheet Manager Module for Stanford Law Review
Handles reading and writing to Excel Master Sheet
"""

import logging
from pathlib import Path
from typing import List, Tuple, Dict, Any, Optional
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink

logger = logging.getLogger(__name__)


class SpreadsheetManager:
    """Manages Excel spreadsheet operations for the SLR system"""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.master_sheet_path = Path(config['paths']['master_sheet'])
        self.sourcepull_sheet = 'Sourcepull'
        self.cc_round1_sheet = 'CC_Round1'
        self.cc_round2_sheet = 'CC_Round2'
        
        # Column mappings for Sourcepull tab
        self.sourcepull_columns = {
            'source_id': 'Source ID',
            'short_name': 'Short Name',
            'citation': 'Full Citation',
            'completed': 'Completed?',
            'location': 'Location',
            'file_name': 'File Name',
            'problems': 'Problems/Comments',
            'retrieved_date': 'Retrieved Date',
            'source_type': 'Source Type'
        }
        
        # Column mappings for CC tabs
        self.cc_columns = {
            'footnote_num': 'Footnote #',
            'cite_order': 'Cite Order',
            'source_id': 'Source ID',
            'quote': 'Quote',
            'supported': 'Supported?',
            'issues': 'Issues',
            'notes': 'Notes',
            'bluebook_compliance': 'Bluebook Compliance',
            'page_found': 'Page Found'
        }
    
    def load_sourcepull_citations(self) -> List[Tuple[str, str, str]]:
        """Load citations from the Sourcepull tab
        Returns: List of tuples (citation_text, source_id, short_name)
        """
        citations = []
        
        try:
            if not self.master_sheet_path.exists():
                logger.error(f"Master sheet not found: {self.master_sheet_path}")
                return citations
            
            # Read the Sourcepull sheet
            df = pd.read_excel(
                self.master_sheet_path,
                sheet_name=self.sourcepull_sheet,
                dtype=str
            )
            
            # Clean column names
            df.columns = df.columns.str.strip()
            
            # Extract citation data
            for idx, row in df.iterrows():
                citation_text = row.get(self.sourcepull_columns['citation'], '')
                source_id = row.get(self.sourcepull_columns['source_id'], '')
                short_name = row.get(self.sourcepull_columns['short_name'], '')
                
                # Skip if already completed
                completed = row.get(self.sourcepull_columns['completed'], '').lower()
                if completed in ['yes', 'y', 'true', '1']:
                    logger.info(f"Skipping completed source: {source_id}")
                    continue
                
                if citation_text and source_id:
                    # Convert source_id to string format with padding
                    try:
                        source_id = f"{int(float(source_id)):03d}"
                    except (ValueError, TypeError):
                        source_id = str(source_id)
                    
                    citations.append((citation_text, source_id, short_name or ''))
            
            logger.info(f"Loaded {len(citations)} citations from spreadsheet")
            
        except Exception as e:
            logger.error(f"Error loading citations: {e}")
        
        return citations
    
    def update_sourcepull_results(self, results: List[Tuple[Any, Any]]) -> bool:
        """Update the Sourcepull tab with retrieval results"""
        try:
            # Load workbook
            wb = openpyxl.load_workbook(self.master_sheet_path)
            ws = wb[self.sourcepull_sheet]
            
            # Find column indices
            header_row = 1
            columns = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value:
                    cell_value = str(cell_value).strip()
                    for key, col_name in self.sourcepull_columns.items():
                        if col_name == cell_value:
                            columns[key] = col
                            break
            
            # Update each result
            for citation, result in results:
                # Find the row for this source
                source_id = citation.source_id
                row_num = self._find_source_row(ws, source_id, columns.get('source_id', 1))
                
                if row_num:
                    # Update completed status
                    if 'completed' in columns:
                        if result.status.value in ['success', 'cached']:
                            ws.cell(row=row_num, column=columns['completed']).value = 'Yes'
                            # Apply green fill
                            ws.cell(row=row_num, column=columns['completed']).fill = PatternFill(
                                start_color="90EE90", end_color="90EE90", fill_type="solid"
                            )
                        else:
                            ws.cell(row=row_num, column=columns['completed']).value = 'No'
                            # Apply red fill for manual required
                            if result.status.value == 'manual_required':
                                ws.cell(row=row_num, column=columns['completed']).fill = PatternFill(
                                    start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                                )
                    
                    # Update file location
                    if 'location' in columns and result.file_path:
                        ws.cell(row=row_num, column=columns['location']).value = result.file_path
                    
                    # Update file name
                    if 'file_name' in columns and result.file_path:
                        file_name = Path(result.file_path).name
                        ws.cell(row=row_num, column=columns['file_name']).value = file_name
                        
                        # Add hyperlink if file exists
                        if Path(result.file_path).exists():
                            cell = ws.cell(row=row_num, column=columns['file_name'])
                            cell.hyperlink = f"file:///{Path(result.file_path).absolute()}"
                            cell.font = Font(color="0000FF", underline="single")
                    
                    # Update problems/comments
                    if 'problems' in columns:
                        if result.message:
                            ws.cell(row=row_num, column=columns['problems']).value = result.message
                    
                    # Update retrieved date
                    if 'retrieved_date' in columns:
                        ws.cell(row=row_num, column=columns['retrieved_date']).value = datetime.now().strftime("%Y-%m-%d")
                    
                    # Update source type
                    if 'source_type' in columns:
                        ws.cell(row=row_num, column=columns['source_type']).value = citation.type.value
            
            # Save workbook
            wb.save(self.master_sheet_path)
            wb.close()
            
            logger.info("Successfully updated Sourcepull results in spreadsheet")
            return True
            
        except Exception as e:
            logger.error(f"Error updating spreadsheet: {e}")
            return False
    
    def load_citechecking_data(self, round_num: int = 1) -> pd.DataFrame:
        """Load citechecking data from CC sheet"""
        sheet_name = self.cc_round1_sheet if round_num == 1 else self.cc_round2_sheet
        
        try:
            df = pd.read_excel(
                self.master_sheet_path,
                sheet_name=sheet_name,
                dtype=str
            )
            return df
        except Exception as e:
            logger.error(f"Error loading CC data: {e}")
            return pd.DataFrame()
    
    def update_citechecking_results(self, results: List[Dict[str, Any]], round_num: int = 1) -> bool:
        """Update citechecking results in the appropriate CC sheet"""
        sheet_name = self.cc_round1_sheet if round_num == 1 else self.cc_round2_sheet
        
        try:
            wb = openpyxl.load_workbook(self.master_sheet_path)
            
            # Create sheet if it doesn't exist
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                self._create_cc_headers(ws)
            else:
                ws = wb[sheet_name]
            
            # Find column indices
            columns = self._get_column_indices(ws, self.cc_columns)
            
            # Update results
            for result in results:
                footnote_num = result.get('footnote_num')
                
                for cite_result in result.get('citations', []):
                    source_id = cite_result.get('source_id')
                    row_num = self._find_cc_row(ws, footnote_num, source_id, columns)
                    
                    if not row_num:
                        # Add new row
                        row_num = ws.max_row + 1
                        ws.cell(row=row_num, column=columns['footnote_num']).value = footnote_num
                        ws.cell(row=row_num, column=columns['source_id']).value = source_id
                    
                    # Update support status
                    if 'supported' in columns:
                        supported = cite_result.get('supported', 'Unknown')
                        ws.cell(row=row_num, column=columns['supported']).value = supported
                        
                        # Apply color coding
                        if supported == 'Yes':
                            fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif supported == 'No':
                            fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                        else:
                            fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                        
                        ws.cell(row=row_num, column=columns['supported']).fill = fill
                    
                    # Update issues
                    if 'issues' in columns and cite_result.get('issues'):
                        ws.cell(row=row_num, column=columns['issues']).value = ', '.join(cite_result['issues'])
                    
                    # Update notes
                    if 'notes' in columns and cite_result.get('notes'):
                        ws.cell(row=row_num, column=columns['notes']).value = cite_result['notes']
                    
                    # Update page found
                    if 'page_found' in columns and cite_result.get('page_found'):
                        ws.cell(row=row_num, column=columns['page_found']).value = str(cite_result['page_found'])
            
            wb.save(self.master_sheet_path)
            wb.close()
            
            logger.info(f"Successfully updated CC Round {round_num} results")
            return True
            
        except Exception as e:
            logger.error(f"Error updating CC results: {e}")
            return False
    
    def update_bluebook_results(self, violations: List[Dict[str, Any]]) -> bool:
        """Update Bluebook compliance results"""
        try:
            wb = openpyxl.load_workbook(self.master_sheet_path)
            
            # Create Bluebook sheet if it doesn't exist
            sheet_name = 'Bluebook_Check'
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                # Add headers
                headers = ['Footnote #', 'Citation', 'Violation Type', 'Issue', 'Suggestion', 'Severity']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col).value = header
                    ws.cell(row=1, column=col).font = Font(bold=True)
            else:
                ws = wb[sheet_name]
            
            # Add violations
            row = ws.max_row + 1
            for violation in violations:
                ws.cell(row=row, column=1).value = violation.get('footnote_num', '')
                ws.cell(row=row, column=2).value = violation.get('citation', '')
                ws.cell(row=row, column=3).value = violation.get('type', '')
                ws.cell(row=row, column=4).value = violation.get('issue', '')
                ws.cell(row=row, column=5).value = violation.get('suggestion', '')
                ws.cell(row=row, column=6).value = violation.get('severity', 'Medium')
                
                # Apply color based on severity
                severity = violation.get('severity', 'Medium')
                if severity == 'High':
                    fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif severity == 'Medium':
                    fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                else:
                    fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = fill
                
                row += 1
            
            wb.save(self.master_sheet_path)
            wb.close()
            
            logger.info("Successfully updated Bluebook results")
            return True
            
        except Exception as e:
            logger.error(f"Error updating Bluebook results: {e}")
            return False
    
    def create_summary_sheet(self) -> bool:
        """Create a summary sheet with overall statistics"""
        try:
            wb = openpyxl.load_workbook(self.master_sheet_path)
            
            # Create or get summary sheet
            sheet_name = 'Summary'
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            ws = wb.create_sheet(sheet_name, 0)  # Insert at beginning
            
            # Add title
            ws.merge_cells('A1:D1')
            ws['A1'] = 'Stanford Law Review Editorial Automation Summary'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Add timestamp
            ws['A3'] = 'Generated:'
            ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Calculate statistics
            stats = self._calculate_statistics(wb)
            
            # Add statistics
            row = 5
            ws[f'A{row}'] = 'Sourcepull Statistics'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for key, value in stats['sourcepull'].items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
            ws[f'A{row}'] = 'Citechecking Statistics'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for key, value in stats['citechecking'].items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            # Format columns
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            
            wb.save(self.master_sheet_path)
            wb.close()
            
            logger.info("Successfully created summary sheet")
            return True
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {e}")
            return False
    
    def _find_source_row(self, ws, source_id: str, source_col: int) -> Optional[int]:
        """Find the row number for a given source ID"""
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=source_col).value
            if cell_value:
                # Handle both string and numeric source IDs
                try:
                    if f"{int(float(cell_value)):03d}" == source_id or str(cell_value) == source_id:
                        return row
                except (ValueError, TypeError):
                    if str(cell_value) == source_id:
                        return row
        return None
    
    def _find_cc_row(self, ws, footnote_num: str, source_id: str, columns: Dict[str, int]) -> Optional[int]:
        """Find the row for a specific footnote and source combination"""
        fn_col = columns.get('footnote_num', 1)
        src_col = columns.get('source_id', 3)
        
        for row in range(2, ws.max_row + 1):
            fn_value = ws.cell(row=row, column=fn_col).value
            src_value = ws.cell(row=row, column=src_col).value
            
            if str(fn_value) == str(footnote_num) and str(src_value) == str(source_id):
                return row
        
        return None
    
    def _get_column_indices(self, ws, column_mapping: Dict[str, str]) -> Dict[str, int]:
        """Get column indices from header row"""
        columns = {}
        header_row = 1
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                cell_value = str(cell_value).strip()
                for key, col_name in column_mapping.items():
                    if col_name == cell_value:
                        columns[key] = col
                        break
        
        return columns
    
    def _create_cc_headers(self, ws):
        """Create headers for CC sheet"""
        headers = list(self.cc_columns.values())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = header
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
    
    def _calculate_statistics(self, wb) -> Dict[str, Dict[str, Any]]:
        """Calculate statistics from all sheets"""
        stats = {
            'sourcepull': {},
            'citechecking': {}
        }
        
        # Sourcepull statistics
        if self.sourcepull_sheet in wb.sheetnames:
            ws = wb[self.sourcepull_sheet]
            total_sources = ws.max_row - 1  # Exclude header
            completed = 0
            
            columns = self._get_column_indices(ws, self.sourcepull_columns)
            if 'completed' in columns:
                for row in range(2, ws.max_row + 1):
                    value = ws.cell(row=row, column=columns['completed']).value
                    if value and str(value).lower() in ['yes', 'y', 'true', '1']:
                        completed += 1
            
            stats['sourcepull'] = {
                'Total Sources': total_sources,
                'Completed': completed,
                'Pending': total_sources - completed,
                'Completion Rate': f"{(completed/total_sources*100):.1f}%" if total_sources > 0 else "0%"
            }
        
        # Citechecking statistics
        for sheet_name in [self.cc_round1_sheet, self.cc_round2_sheet]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                total_cites = ws.max_row - 1
                
                columns = self._get_column_indices(ws, self.cc_columns)
                if 'supported' in columns:
                    supported = 0
                    unsupported = 0
                    
                    for row in range(2, ws.max_row + 1):
                        value = ws.cell(row=row, column=columns['supported']).value
                        if value:
                            if str(value).lower() in ['yes', 'y', 'true', '1']:
                                supported += 1
                            elif str(value).lower() in ['no', 'n', 'false', '0']:
                                unsupported += 1
                    
                    round_name = 'Round 1' if 'Round1' in sheet_name else 'Round 2'
                    stats['citechecking'][f'{round_name} Total'] = total_cites
                    stats['citechecking'][f'{round_name} Supported'] = supported
                    stats['citechecking'][f'{round_name} Not Supported'] = unsupported
        
        return stats
    
    def export_to_csv(self, output_dir: str = "exports") -> bool:
        """Export all sheets to CSV files"""
        try:
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)
            
            # Read all sheets
            xl_file = pd.ExcelFile(self.master_sheet_path)
            
            for sheet_name in xl_file.sheet_names:
                df = pd.read_excel(xl_file, sheet_name=sheet_name)
                csv_path = output_path / f"{sheet_name}.csv"
                df.to_csv(csv_path, index=False)
                logger.info(f"Exported {sheet_name} to {csv_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return False