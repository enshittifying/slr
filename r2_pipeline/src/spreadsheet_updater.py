"""
Update Excel spreadsheet with R2 results.
"""
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, List, Optional
import logging

logger = logging.getLogger(__name__)

class SpreadsheetUpdater:
    """Update Excel spreadsheet with R2 data."""

    def __init__(self, spreadsheet_path: Path):
        self.spreadsheet_path = spreadsheet_path

        if not spreadsheet_path.exists():
            logger.warning(f"Spreadsheet not found at {spreadsheet_path}. Creating mock updater.")
            self.wb = None
            self.sheet = None
            return

        self.wb = load_workbook(spreadsheet_path)

        # Identify the correct sheet (assumes "CC (nn. 78-113); HC" or similar)
        self.sheet = None
        for sheet_name in self.wb.sheetnames:
            if "CC" in sheet_name and "78" in sheet_name:
                self.sheet = self.wb[sheet_name]
                break

        if self.sheet is None:
            logger.warning("Could not find appropriate CC sheet in workbook")

        logger.info(f"Using sheet: {self.sheet.title if self.sheet else 'None'}")

    def update_citation(self,
                       footnote_num: int,
                       cite_num: int,
                       results: Dict) -> None:
        """
        Update a single citation's R2 data.

        Args:
            footnote_num: Footnote number
            cite_num: Citation number within footnote
            results: Dict with validation results
        """
        if not self.sheet:
            logger.warning("No sheet available for update")
            return

        # Find the row for this citation
        row = self._find_citation_row(footnote_num, cite_num)

        if row is None:
            logger.warning(f"Could not find row for fn {footnote_num}, cite {cite_num}")
            return

        # Determine column indices (adjust based on actual spreadsheet structure)
        # From your spreadsheet: columns are like "Fn#", "Cite#", "Supports?", "Citation Elements", "MEM Comments"
        # R2 columns start after R1 columns

        # Update R2 columns
        r2_fn_col = self._find_column("Fn#", start_col=17)  # R2 section
        r2_cite_col = self._find_column("Cite#", start_col=17)
        r2_supports_col = self._find_column("Supports?", start_col=17)
        r2_elements_col = self._find_column("Citation Elements", start_col=17)
        r2_comments_col = self._find_column("MEM Comments", start_col=17)

        # Populate cells
        if r2_fn_col:
            self.sheet.cell(row=row, column=r2_fn_col, value=footnote_num)
        if r2_cite_col:
            self.sheet.cell(row=row, column=r2_cite_col, value=cite_num)

        # Supports?
        if r2_supports_col:
            support = self._determine_support_value(results)
            self.sheet.cell(row=row, column=r2_supports_col, value=support)

        # Citation Elements
        if r2_elements_col:
            elements = self._format_citation_elements(results)
            self.sheet.cell(row=row, column=r2_elements_col, value=elements)

        # MEM Comments
        if r2_comments_col:
            comments = self._format_mem_comments(results)
            self.sheet.cell(row=row, column=r2_comments_col, value=comments)

        logger.info(f"Updated spreadsheet row {row} for fn {footnote_num}, cite {cite_num}")

    def _find_citation_row(self, footnote_num: int, cite_num: int) -> Optional[int]:
        """Find the row number for a specific citation."""
        # Search through rows to find matching footnote and cite numbers
        fn_col = self._find_column("Fn#")
        cite_col = self._find_column("Cite#")

        if not fn_col or not cite_col:
            return None

        for row in range(4, self.sheet.max_row + 1):  # Start from row 4 (after headers)
            fn_val = self.sheet.cell(row=row, column=fn_col).value
            cite_val = self.sheet.cell(row=row, column=cite_col).value

            if fn_val == footnote_num and cite_val == cite_num:
                return row

        return None

    def _find_column(self, header_name: str, start_col: int = 1) -> Optional[int]:
        """Find column number by header name."""
        for col in range(start_col, self.sheet.max_column + 1):
            header = self.sheet.cell(row=2, column=col).value  # Row 2 has headers
            if header and header_name.lower() in str(header).lower():
                return col
        return None

    def _determine_support_value(self, results: Dict) -> str:
        """Determine "Supports?" value from results."""
        if "support_analysis" in results and results["support_analysis"]:
            support_level = results["support_analysis"].get("support_level", "").lower()
            if support_level == "yes":
                return "Yes"
            elif support_level == "maybe":
                return "Maybe"
            elif support_level == "no":
                return "No"
        return "To Check"

    def _format_citation_elements(self, results: Dict) -> str:
        """Format citation elements for spreadsheet."""
        if "citation_validation" in results and results["citation_validation"]:
            cv = results["citation_validation"]
            if cv.get("is_correct"):
                return "No Issues"
            else:
                errors = cv.get("errors", [])
                if errors:
                    error_types = [e.get("error_type", "unknown") for e in errors]
                    return f"Repaired Issue - {', '.join(error_types)}"
        return "No Issues"

    def _format_mem_comments(self, results: Dict) -> str:
        """Format MEM comments from results."""
        comments = []

        # Citation validation
        if "citation_validation" in results and results["citation_validation"]:
            cv = results["citation_validation"]
            if not cv.get("is_correct"):
                errors = cv.get("errors", [])
                for error in errors[:2]:  # First 2 errors
                    comments.append(error.get("description", ""))

        # Support analysis
        if "support_analysis" in results and results["support_analysis"]:
            sa = results["support_analysis"]
            reasoning = sa.get("reasoning", "")[:100]  # First 100 chars
            comments.append(f"Support: {reasoning}")

        # Quote verification
        if "quote_verification" in results and results["quote_verification"]:
            qv = results["quote_verification"]
            if not qv.get("accurate"):
                comments.append(f"Quote issues: {len(qv.get('issues', []))} found")

        # Overall recommendation
        if "recommendation" in results:
            comments.append(f"Action: {results['recommendation']}")

        return "; ".join(comments)

    def save(self) -> None:
        """Save the workbook."""
        if self.wb:
            self.wb.save(self.spreadsheet_path)
            logger.info(f"Saved spreadsheet to {self.spreadsheet_path}")

    def close(self):
        """Close the workbook."""
        if self.wb:
            self.wb.close()
