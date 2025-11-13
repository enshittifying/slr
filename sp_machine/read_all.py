from openpyxl import load_workbook
from pathlib import Path

wb = load_workbook(Path(__file__).parent.parent / "78 SLR V2 R2 F" / "References" / "V78.4 Bersh Master Sheet.xlsx")
ws = wb.active

for row in ws.iter_rows():
    print([cell.value for cell in row])
